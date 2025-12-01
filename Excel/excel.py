# в первую очередь устанавливаем библиотеку openpyxl
import openpyxl

# открываем файл для чтения
book = openpyxl.open('грузчики црм, артём.xlsx', read_only=True)

# по умолчанию первая страница active
sheet = book.active
# можно обращаться по строке и столбцу, либо по строковому названию ячейки
# print(sheet[2][0].value)
# print(sheet['A2'].value)

# пробегаем циклом по всем строкам и столбцам с помощью встроенного модуля max_row
# for row in range(1, int(abs(sheet.max_row/26))): - такую конструкцию делал чтобы можно было делить,
# т.к. при делении возникает float который выдаёт ошибку
# for row in range(1, sheet.max_row+1):
#     phone_number = sheet[row][0].value
#     name_loader = sheet[row][1].value
#     company_loader = sheet[row][2].value
#     print(row, phone_number, name_loader, company_loader)

# обращение к листам по индексу
# sheet_1 = book.worksheets[0]
# print(sheet_1[2][2].value)

# обращение по диапазону
# cells = sheet['A3':'C7']
# диапазон выводится картеж картежей из 3-х значений ((A1, B1, C1), (A2, B2, C2))
# print(cells)
# for cell in cells:
#     print(cell)
# обходя циклом можно распаковать значения картежа, нужно присвоить 3 переменных
# for phone, name, company  in cells:
#     print(phone.value, name.value, company.value)

# метод iter_rows (min_row-начало рядов, max_row-конец рядов, min_col-начало столбцов, max_col-количество столбцов)
# если не передавать атрибуты в sheet.iter_rows() то он обойдёт всю страницу
for row in sheet.iter_rows(min_row=2, max_row=10, min_col=0, max_col=3):
    for cell in row:
        # end=' ', чтобы не писать каждый cell с новой строки ставим пробел
        print(cell.value, end=' ')
    # по окончании ряда(row) перевод на новую строку
    print()